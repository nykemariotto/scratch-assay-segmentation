# -*- coding: utf-8 -*-
"""
stage1/make_annotation_xlsx.py — converte stage1/annotation_sheet.csv em stage1/annotation_sheet.xlsx
para preenchimento manual confortavel.

Recursos:
  - aba "Instrucoes" com legenda das colunas + LINHA DE EXEMPLO preenchida
  - aba "Anotacao" com as 158 imagens, ordenadas por serie
  - painel congelado: contexto (ordem..area) fixo, colunas de preenchimento a vista
  - autofilter, larguras ajustadas, fonte Arial
  - listas suspensas: pasta_correta? (sim/nao/duvida), confianca (alta/media/baixa)
  - categoria_sugerida: sugestoes na lista, mas ACEITA texto novo (a ideia e
    justamente descobrir categorias que nao existem ainda)
  - celulas a preencher em amarelo; contexto em cinza claro
  - contador de progresso na aba Instrucoes (formula COUNTA)

NAO mostra 'modo_falha' nem 'contencao_t0': a justificativa precisa ser
independente da metrica geometrica para servir de validacao dela.
"""
import csv, os, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

SRC = "stage1/annotation_sheet.csv"
OUT = "stage1/annotation_sheet.xlsx"
if not os.path.isfile(SRC):
    sys.exit(f"did not find {SRC} (run stage1/build_annotation_sheet.py)")

rows = list(csv.DictReader(open(SRC, encoding="utf-8-sig")))

# Column order: the EXACT file name sits immediately before the fill-in columns
# and inside the frozen pane, so that it is always visible, and it is a clickable
# link that opens the image directly, which avoids searching by a similar name.
COLS = [
    ("ordem", 6, "ctx"), ("pasta_triagem", 18, "ctx"),
    ("arquivo_overlay", 58, "ctx"),
    ("motivo", 50, "fill"), ("pasta_correta?", 14, "fill"),
    ("categoria_sugerida", 26, "fill"), ("confianca", 11, "fill"),
    ("analysis_unit", 34, "ref"), ("timepoint_h", 7, "ref"), ("campo", 7, "ref"),
    ("area_pct_whst", 11, "ref"), ("cell_line", 10, "ref"),
    ("adjudicado_depois_como", 20, "ref"), ("whst_input_file", 58, "ref"),
]
# pasta_triagem -> subpasta real em whst_output/overlays_sorted/
SUBDIR = {"_SEG_RUIM/_super": "_SEG_RUIM/_super", "_SEG_RUIM/_sub": "_SEG_RUIM/_sub",
          "_IMG_INVALIDA": "_IMG_INVALIDA", "_AMBIGUO": "_AMBIGUO"}
# letras de coluna derivadas de COLS (reordenar COLS nao quebra formulas/validacoes)
from openpyxl.utils import get_column_letter as _gcl
L = {name: _gcl(j) for j, (name, _, _) in enumerate(COLS, start=1)}

# ---- preserva anotacoes ja feitas (chave: whst_input_file) ----
ANT = {}
if os.path.isfile(OUT):
    from openpyxl import load_workbook
    _ws = load_workbook(OUT, data_only=True)["Anotacao"]
    _h = [c.value for c in _ws[1]]
    if "whst_input_file" in _h:
        _ci = {h: j for j, h in enumerate(_h, 1)}
        for _i in range(2, _ws.max_row + 1):
            _k = _ws.cell(_i, _ci["whst_input_file"]).value
            if not _k:
                continue
            d = {c: _ws.cell(_i, _ci[c]).value for c in
                 ("motivo", "pasta_correta?", "categoria_sugerida", "confianca") if c in _ci}
            if any(v not in (None, "") for v in d.values()):
                ANT[str(_k)] = {k2: ("" if v is None else v) for k2, v in d.items()}
    print(f"preservando {len(ANT)} anotacoes ja feitas")
ARIAL = "Arial"
YEL = PatternFill("solid", fgColor="FFF2CC")     # a preencher
GRY = PatternFill("solid", fgColor="F2F2F2")     # contexto
HDR = PatternFill("solid", fgColor="1F4E79")
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

wb = Workbook()

# ---------------------------------------------------------------- Instrucoes
ins = wb.active
ins.title = "Instrucoes"
ins.sheet_view.showGridLines = False
A = lambda r, c, v, **kw: _put(ins, r, c, v, **kw)


def _put(ws, r, c, v, bold=False, size=11, color="000000", fill=None,
         wrap=False, italic=False):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(name=ARIAL, bold=bold, size=size, color=color, italic=italic)
    if fill:
        cell.fill = fill
    cell.alignment = Alignment(vertical="top", wrap_text=wrap)
    return cell


A(1, 1, "Anotação da triagem visual — por que cada imagem foi para aquela pasta",
  bold=True, size=15, color="1F4E79")
A(2, 1, "Manuscrito 4336348 · Cytometry Part A", size=10, color="595959")

A(4, 1, "O QUE PREENCHER (aba \"Anotacao\", colunas em AMARELO)", bold=True, size=12)
leg = [
    ("motivo", "Por que você colocou essa imagem nessa pasta. Texto livre, pode ser curto."),
    ("pasta_correta?", "Revendo agora: a pasta estava certa?  sim / nao / duvida"),
    ("categoria_sugerida",
     "Se 'nao' ou 'duvida': que categoria descreveria melhor. PODE INVENTAR NOMES — "
     "descobrir categorias que faltam é justamente o objetivo. A lista suspensa traz "
     "sugestões, mas você pode digitar qualquer coisa."),
    ("confianca", "Opcional: alta / media / baixa"),
]
r = 5
for nome, desc in leg:
    A(r, 1, nome, bold=True, fill=YEL)
    A(r, 2, desc, wrap=True)
    ins.row_dimensions[r].height = 30
    r += 1

r += 1
A(r, 1, "COLUNAS DE CONTEXTO (cinza) — não editar", bold=True, size=12); r += 1
for nome, desc in [
    ("pasta_triagem", "a pasta em que você colocou durante a triagem cega"),
    ("analysis_unit / timepoint_h / campo", "identificação da série — permite ler o contexto temporal"),
    ("area_pct_whst", "área que o WHST mediu (a mesma que aparecia no overlay)"),
    ("arquivo_overlay", "nome do arquivo em whst_output/overlays_sorted/ para localizar a imagem"),
    ("adjudicado_depois_como", "só para as 12 que eram AMBIGUO: como foram adjudicadas depois"),
]:
    A(r, 1, nome, bold=True, fill=GRY); A(r, 2, desc, wrap=True); r += 1

r += 1
A(r, 1, "EXEMPLO DE PREENCHIMENTO", bold=True, size=12); r += 1
ex_hdr = ["pasta_triagem", "analysis_unit", "tp", "motivo", "pasta_correta?",
          "categoria_sugerida", "confianca"]
for j, h in enumerate(ex_hdr, start=1):
    c = A(r, j, h, bold=True, color="FFFFFF"); c.fill = HDR; c.border = BOX
r += 1
ex = ["_SEG_RUIM/_super", "originais||None||D4", 24,
      "ferida já fechada; o contorno pegou um aglomerado de debris no canto",
      "nao", "pegou_debris_apos_fechamento", "alta"]
for j, v in enumerate(ex, start=1):
    c = A(r, j, v, italic=True)
    c.fill = YEL if j >= 4 else GRY
    c.border = BOX
ins.row_dimensions[r].height = 28
ex_row = r

r += 2
A(r, 1, "PROGRESSO", bold=True, size=12); r += 1
_ult = len(rows) + 1
A(r, 1, "linhas com 'motivo' preenchido:", bold=True)
ins.cell(row=r, column=3,
         value=f"=COUNTA(Anotacao!{L['motivo']}2:{L['motivo']}{_ult})"
         ).font = Font(name=ARIAL, bold=True, size=12)
r += 1
A(r, 1, "total de linhas:", bold=True)
ins.cell(row=r, column=3, value=f"=COUNTA(Anotacao!A2:A{_ult})").font = Font(name=ARIAL, size=11)
r += 2
A(r, 1, "Não precisa preencher tudo de uma vez, nem todas as linhas — "
        "linhas em branco são permitidas.", italic=True, color="595959", wrap=True)
r += 1
A(r, 1, "Ao terminar (ou parcialmente):  python stage1/read_annotations.py",
  bold=True, color="1F4E79")

for col, wdt in (("A", 34), ("B", 62), ("C", 14), ("D", 46), ("E", 15), ("F", 28), ("G", 12)):
    ins.column_dimensions[col].width = wdt

# ---------------------------------------------------------------- Anotacao
ws = wb.create_sheet("Anotacao")
for j, (name, wdt, _) in enumerate(COLS, start=1):
    c = ws.cell(row=1, column=j, value=name)
    c.font = Font(name=ARIAL, bold=True, size=11, color="FFFFFF")
    c.fill = HDR
    c.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    c.border = BOX
    ws.column_dimensions[get_column_letter(j)].width = wdt
ws.row_dimensions[1].height = 30

THICK = Side(style="medium", color="1F4E79")
prev_pasta = None
for i, r0 in enumerate(rows, start=2):
    nova_pasta = r0["pasta_triagem"] != prev_pasta      # inicio de bloco
    prev_pasta = r0["pasta_triagem"]
    ant = ANT.get(r0["whst_input_file"], {})
    for j, (name, _, kind) in enumerate(COLS, start=1):
        v = ant.get(name, r0.get(name, "")) if kind == "fill" else r0.get(name, "")
        if name in ("ordem", "timepoint_h"):
            v = int(v) if str(v).strip() else None
        elif name == "area_pct_whst":
            v = float(v) if str(v).strip() else None
        c = ws.cell(row=i, column=j, value=v)
        c.font = Font(name=ARIAL, size=10,
                      bold=(nova_pasta and name == "pasta_triagem"))
        # nome do arquivo = link que abre a imagem direto
        if name == "arquivo_overlay" and v:
            sub = SUBDIR.get(r0["pasta_triagem"], "")
            c.hyperlink = f"whst_output/overlays_sorted/{sub}/{v}"
            c.font = Font(name=ARIAL, size=10, color="0563C1", underline="single")
        # borda grossa no topo separa visualmente os blocos de pasta
        c.border = Border(left=THIN, right=THIN, bottom=THIN,
                          top=(THICK if nova_pasta else THIN))
        c.alignment = Alignment(vertical="top", wrap_text=(name == "motivo"))
        if kind == "fill":
            c.fill = YEL
        elif kind == "ctx":
            c.fill = GRY
    ws.row_dimensions[i].height = 15

last = len(rows) + 1
# letras derivadas de COLS -> reordenar as colunas nao quebra nada abaixo
L = {name: get_column_letter(j) for j, (name, _, _) in enumerate(COLS, start=1)}
# congela ate a coluna anterior a primeira de preenchimento: nome do arquivo
# fica sempre visivel enquanto se digita
primeira_fill = next(j for j, (_, _, k) in enumerate(COLS, start=1) if k == "fill")
ws.freeze_panes = f"{get_column_letter(primeira_fill)}2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{last}"
col_area = COLS.index(("area_pct_whst", 11, "ref")) + 1
for i in range(2, last + 1):
    ws.cell(row=i, column=col_area).number_format = "0.000"

# ---- listas suspensas ----
def dv_list(items, allow_other=False):
    d = DataValidation(type="list", formula1='"' + ",".join(items) + '"',
                       allow_blank=True, showErrorMessage=not allow_other)
    if allow_other:
        d.error = None
    return d


dv_ok = dv_list(["sim", "nao", "duvida"])
dv_ok.prompt = "sim / nao / duvida"
dv_ok.promptTitle = "A pasta estava certa?"
ws.add_data_validation(dv_ok)
dv_ok.add(f"{L['pasta_correta?']}2:{L['pasta_correta?']}{last}")

dv_cf = dv_list(["alta", "media", "baixa"])
ws.add_data_validation(dv_cf)
dv_cf.add(f"{L['confianca']}2:{L['confianca']}{last}")

# categoria_sugerida: sugestoes, mas ACEITA texto novo (showErrorMessage=False)
SUG = ["pegou_fundo_do_poco", "pegou_debris", "contorno_fora_da_ferida",
       "ferida_ja_fechada", "dividiu_a_ferida", "pegou_so_parte_da_ferida",
       "borda_do_poco", "sem_ferida_no_campo", "monolayer_nao_confluente"]
dv_sug = dv_list(SUG, allow_other=True)
dv_sug.prompt = "Escolha uma sugestão OU digite uma categoria nova"
dv_sug.promptTitle = "categoria_sugerida"
ws.add_data_validation(dv_sug)
dv_sug.add(f"{L['categoria_sugerida']}2:{L['categoria_sugerida']}{last}")

wb.save(OUT)
print(f"gerado: {OUT}")
print(f"  aba 'Anotacao': {len(rows)} linhas x {len(COLS)} colunas")
print(f"  painel congelado em {ws.freeze_panes} -> colunas A..{get_column_letter(primeira_fill-1)} "
      f"sempre visiveis (inclui o NOME DO ARQUIVO em {L['arquivo_overlay']})")
print(f"  the file name is a HYPERLINK: clicking opens the image directly")
print(f"  listas suspensas: {L['pasta_correta?']} (sim/nao/duvida), "
      f"{L['confianca']} (alta/media/baixa), {L['categoria_sugerida']} (sugestoes + livre)")
print(f"  anotacoes preservadas: {len(ANT)}")
print(f"  aba 'Instrucoes': legenda + linha de exemplo (linha {ex_row}) + contador de progresso")
