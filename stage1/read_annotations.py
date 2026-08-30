# -*- coding: utf-8 -*-
"""
stage1/read_annotations.py — le stage1/annotation_sheet.csv preenchido e produz tres analises:

  (1) TRIAGE AUDIT
      How many images the operator, on review with a justification, considers to
      have been in the wrong folder. That is an error rate of the triage itself —
      reportable, and more honest than assuming the triage was perfect.

  (2) CATEGORIAS NOVAS
      Agrupa 'categoria_sugerida' e minera os 'motivo' por palavras recorrentes.
      If several images describe the same situation without fitting the original
      4 boxes, that is a category missing from the scheme.

  (3) VALIDACAO INDEPENDENTE DO 'modo_falha'
      Crosses the operator's justification (given without seeing the metric) with the
      modo_falha computado por contencao geometrica. Concordancia = evidencia
      convergent; disagreement localises an error in the metric or in the triage.

Nada e alterado automaticamente: o script so relata. Reclassificar exige
decisao explicita depois.
"""
import csv, os, re, sys
from collections import Counter, defaultdict

SHEET_CSV = "stage1/annotation_sheet.csv"
SHEET_XLSX = "stage1/annotation_sheet.xlsx"
ATUAL = "data/visual_triage.csv"
OUT = "data/annotation_report.csv"

CAMPOS = ["motivo", "pasta_correta?", "categoria_sugerida", "confianca"]


def ler_xlsx(path):
    """reads the 'Anotacao' tab of the filled-in xlsx.

    The stopword list and the keyword matching below stay in Portuguese on
    purpose: they mine the operator's free text, which was written in Portuguese.
    """
    from openpyxl import load_workbook
    ws = load_workbook(path, data_only=True)["Anotacao"]
    hdr = [c.value for c in ws[1]]
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        d = {h: ("" if v is None else str(v).strip()) for h, v in zip(hdr, row)}
        out.append(d)
    return out


# usa o xlsx se existir e for o mais recente (e o formato de preenchimento)
if os.path.isfile(SHEET_XLSX) and (not os.path.isfile(SHEET_CSV)
                                   or os.path.getmtime(SHEET_XLSX) >= os.path.getmtime(SHEET_CSV)):
    rows = ler_xlsx(SHEET_XLSX)
    print(f"[fonte: {SHEET_XLSX}]")
elif os.path.isfile(SHEET_CSV):
    rows = list(csv.DictReader(open(SHEET_CSV, encoding="utf-8-sig")))
    print(f"[fonte: {SHEET_CSV}]")
else:
    sys.exit("found neither stage1/annotation_sheet.xlsx nor .csv (run stage1/build_annotation_sheet.py)")
for r in rows:                       # garante as colunas de preenchimento
    for c in CAMPOS:
        r.setdefault(c, "")
atual = {r["whst_input_file"]: r for r in csv.DictReader(open(ATUAL, encoding="utf-8-sig"))}

pre = [r for r in rows if r["motivo"].strip() or r["pasta_correta?"].strip()
       or r["categoria_sugerida"].strip()]
print(f"=== COVERAGE ===")
print(f"  rows in the spreadsheet   : {len(rows)}")
print(f"  rows with any annotation: {len(pre)}  ({len(pre)/max(len(rows),1):.0%})")
if not pre:
    sys.exit("\nNo annotation filled in yet. Fill the 'Anotacao' tab of "
             "stage1/annotation_sheet.xlsx (colunas amarelas).")

# ---------------- (1) auditoria ----------------
print("\n=== (1) TRIAGE AUDIT ===")
vc = Counter(r["pasta_correta?"].strip().lower() for r in pre if r["pasta_correta?"].strip())
tot = sum(vc.values())
for k in ("sim", "nao", "duvida"):
    if vc.get(k):
        print(f"  pasta_correta? = {k:<7} {vc[k]:>4}  ({vc[k]/tot:.0%})")
outros = {k: v for k, v in vc.items() if k not in ("sim", "nao", "duvida")}
if outros:
    print(f"  [warning] values outside the vocabulary: {outros}")
err = [r for r in pre if r["pasta_correta?"].strip().lower() in ("nao", "duvida")]
if err:
    print(f"\n  images possibly in the wrong folder: {len(err)}")
    by = Counter(r["pasta_triagem"] for r in err)
    for p, n in by.most_common():
        base = sum(1 for r in rows if r["pasta_triagem"] == p)
        print(f"    {p:<20} {n:>3} de {base:>3} ({n/base:.0%})")

# ---------------- (2) categorias novas ----------------
print("\n=== (2) CATEGORIAS SUGERIDAS ===")
sug = Counter(r["categoria_sugerida"].strip().lower()
              for r in pre if r["categoria_sugerida"].strip())
if sug:
    for c, n in sug.most_common():
        print(f"  {c:<34} {n:>3}")
    fortes = [c for c, n in sug.items() if n >= 3]
    if fortes:
        print(f"\n  >>> candidatas a CATEGORIA NOVA (>=3 ocorrencias): {fortes}")
else:
    print("  no suggestion recorded")

# mining the free-text reasons
STOP = set("""a o as os de do da dos das e em no na nos nas um uma que com por para
se ja mas nao muito pouco ta esta este essa esse isso ele ela foi ser sao tem
mais menos so pois porque como onde quando qual quais ao aos à às""".split())
print("\n  recurring terms in 'motivo' (>=3):")
w = Counter()
for r in pre:
    for t in re.findall(r"[a-zA-ZÀ-ÿ_]{4,}", r["motivo"].lower()):
        if t not in STOP:
            w[t] += 1
for t, n in w.most_common(25):
    if n >= 3:
        print(f"    {t:<26} {n:>3}")

# ---------------- (3) validacao do modo_falha ----------------
print("\n=== (3) CROSS-TABULATION with 'modo_falha' (computed by containment) ===")
print("  the operator annotated WITHOUT seeing this metric -> independent validation")
cross = defaultdict(Counter)
for r in pre:
    k = r["whst_input_file"]
    mf = atual.get(k, {}).get("modo_falha", "") or "(sem)"
    key = (r["categoria_sugerida"].strip().lower() or r["pasta_triagem"])
    cross[mf][key] += 1
for mf in sorted(cross):
    tot = sum(cross[mf].values())
    itens = ", ".join(f"{k}={v}" for k, v in cross[mf].most_common(5))
    print(f"  {mf:<18} n={tot:<4} {itens}")

# ---- classificacao do texto livre em modo de falha ----
# NOTE the distinction that breaks a naive regex:
#   "caught the wound and outside it TOO" -> EXCESS  (caught the wound + extra)
#   "caught outside the wound"           -> DISPLACED (it did not catch the wound)
# Regras por frase, aplicadas em ordem de especificidade.
def classifica_motivo(m):
    m = m.lower().strip()
    if not m:
        return ""
    # 1) out of scope: an algorithm test image, not experimental data
    if "cruz" in m and "teste" in m:
        return "fora_do_escopo"
    # 2) imagem inutilizavel
    if ("invalid" in m) or ("muito ruidosa" in m and "segmenta" not in m) \
       or ("ruidosa" in m and "dificil" in m):
        return "imagem_invalida"
    # 3) closed wound
    if "fechad" in m and "pouco super" not in m:
        return "espuria_fechada"
    # 4) excesso + falta simultaneos (misto)
    if ("e fora dela" in m or "e fora dela tambem" in m) and "faltou" in m:
        return "excesso_e_sub"
    # 5) excess: caught the wound and something beyond
    if "e fora dela" in m or "pouco super" in m:
        return "excesso"
    # 6) displaced: caught outside / wrong place, without catching the wound
    if ("pegou fora" in m or "local errado" in m or "fora do risco" in m
            or "fora da central" in m):
        return "deslocada"
    # 7) sub: pegou so parte
    if "faltou" in m or "apenas uma parte" in m:
        return "sub"
    # 8) segmentacao considerada correta
    if "corret" in m or "segmenta" in m and "boa" in m:
        return "segmentacao_ok"
    # 9) poor image but recoverable
    if "daria para segmentar" in m or "possivel de segmentar" in m or "dificultaria" in m:
        return "ruidosa_recuperavel"
    return "outro"


for r in rows:
    r["motivo_classificado"] = classifica_motivo(r.get("motivo", ""))

print("\n=== (3a) CLASSIFICACAO DO TEXTO LIVRE ===")
cm = Counter(r["motivo_classificado"] for r in pre if r["motivo_classificado"])
for m, n in cm.most_common():
    print(f"  {m:<22} {n:>4}")

# cross-tabulate only where BOTH describe the same axis (excess/shifted/spurious-closed)
COMP = {"excesso": "excesso", "deslocada": "deslocada", "espuria_fechada": "espuria_fechada"}
sinal = []
for r in pre:
    s = COMP.get(r["motivo_classificado"])
    mf = atual.get(r["whst_input_file"], {}).get("modo_falha", "")
    if s and mf in ("excesso", "deslocada", "espuria_fechada"):
        sinal.append((s, mf))
if sinal:
    ok = sum(1 for s, mf in sinal if s == mf)
    print(f"\n=== (3b) CONCORDANCIA texto-livre x metrica geometrica ===")
    print(f"  comparable (both on the excesso/deslocada/fechada axis): {len(sinal)}")
    print(f"  concordam: {ok}/{len(sinal)} ({ok/len(sinal):.0%})")
    mat = Counter(sinal)
    print(f"  {'operador':<18}{'metrica':<18}{'n':>5}")
    for (s, mf), n in mat.most_common():
        flag = "  <-- concorda" if s == mf else ""
        print(f"  {s:<18}{mf:<18}{n:>5}{flag}")

# ---------------- saida ----------------
for r in rows:
    k = r["whst_input_file"]
    r["modo_falha_computado"] = atual.get(k, {}).get("modo_falha", "")
    r["contencao_t0"] = atual.get(k, {}).get("contencao_t0", "")
    r.setdefault("motivo_classificado", classifica_motivo(r.get("motivo", "")))
with open(OUT, "w", encoding="utf-8-sig", newline="") as f:
    w2 = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
    w2.writeheader(); w2.writerows(rows)
print(f"\nSaved: {OUT} (the sheet + modo_falha/contencao for checking)")
